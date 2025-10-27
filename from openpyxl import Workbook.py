from openpyxl import Workbook
# -- cod    
wb = Workbook()
ws = wb.active
ws.title = "Estoque"
ws2 = wb.create_sheet("Cadastro")
ws3 = wb.create_sheet("Lojas")
wb.save("estoques.xlsx")

ws = wb.active
ws.append(["Produto", "Quantidade", "Preço"])
ws.append(["Café", 10, 25])
ws.append(["Açúcar", 5, 10])

ws["D1"] = "Total"

wb.save("/Users/luskoliveira/Documents/DOCUMENTOS FUCAPE/CD_06_10/estoques.xlsx")

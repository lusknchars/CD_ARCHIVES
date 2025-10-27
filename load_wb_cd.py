from openpyxl import load_workbook

wb = load_workbook(
    "/Users/luskoliveira/Documents/DOCUMENTOS FUCAPE/CD_06_10/estoques.xlsx")
ws = wb["Estoque"]

print(ws["A2"].value, ws["B2"].value, ws["C2"].value)
ws["B2"] = 15
for linha in ws.iter_rows(values_only=True):
    print(linha)
wb.save("/Users/luskoliveira/Documents/DOCUMENTOS FUCAPE/CD_06_10/estoques.xlsx")
